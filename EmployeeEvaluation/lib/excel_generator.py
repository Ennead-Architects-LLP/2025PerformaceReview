"""
Excel Report Generator

This module handles the generation of Excel files from employee evaluation data.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict


def create_excel_output(employees: List[Dict[str, str]], all_fields: List[str], output_filename: str = 'OUTPUT.xlsx') -> str:
    """
    Create Excel file with employee data and professional formatting.
    
    Args:
        employees: List of employee dictionaries
        all_fields: List of all field names
        output_filename: Name of the output Excel file
        
    Returns:
        Path to the created Excel file
    """
    # Prepare data for DataFrame with logical column ordering
    excel_data = []
    for emp in employees:
        row = {}
        
        # Define logical column order
        priority_fields = [
            'employee_name',
            'employee_role', 
            'date_of_evaluation',
            'evaluator_name'
        ]
        
        # Add priority fields first
        for field in priority_fields:
            if field == 'employee_name':
                row['Employee Name'] = emp.get('employee_name', '')
            else:
                row[field] = emp.get(field, '')
        
        # Add all other fields in order (excluding already added ones)
        for field in all_fields:
            if field not in priority_fields:
                row[field] = emp.get(field, '')
        
        excel_data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Evaluations"
    
    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Apply professional formatting
    apply_excel_formatting(ws, len(employees), len(df.columns))
    
    # Save the workbook in docs folder for GitHub Pages
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    docs_dir = os.path.join(project_root, 'docs')
    output_path = os.path.join(docs_dir, f"_EvaluationSummary.xlsx")

    # Ensure docs directory exists
    os.makedirs(docs_dir, exist_ok=True)

    wb.save(output_path)

    print(f"Excel file created: {output_path}")
    return output_path


def apply_excel_formatting(ws, num_rows: int, num_cols: int) -> None:
    """
    Apply professional formatting to the Excel worksheet.
    
    Args:
        ws: Worksheet object
        num_rows: Number of data rows
        num_cols: Number of columns
    """
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    data_font = Font(size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Format header row
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Format data rows
    for row in range(2, num_rows + 2):  # +2 because row 1 is header and we start from row 2
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = border
            
            # Apply different alignment based on column
            if col == 1:  # Name column
                cell.alignment = left_alignment
            elif col == 2:  # Time column
                cell.alignment = center_alignment
            else:  # Other fields
                cell.alignment = left_alignment
    
    # Auto-size columns
    for col in range(1, num_cols + 1):
        column_letter = ws.cell(row=1, column=col).column_letter
        max_length = 0
        
        # Check header length
        header_length = len(str(ws.cell(row=1, column=col).value))
        max_length = max(max_length, header_length)
        
        # Check data lengths
        for row in range(2, num_rows + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                # For text wrapping, we need to consider multiple lines
                cell_length = len(str(cell_value))
                # If text is long, we'll set a reasonable width
                if cell_length > 50:
                    cell_length = 50
                max_length = max(max_length, cell_length)
        
        # Set column width with some padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set row heights for better text wrapping
    for row in range(1, num_rows + 2):
        ws.row_dimensions[row].height = 25  # Increased height for text wrapping
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add filters to header row
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=num_cols).coordinate}"


def create_summary_statistics(employees: List[Dict[str, str]], all_fields: List[str]) -> Dict[str, any]:
    """
    Create summary statistics for the employee data.
    
    Args:
        employees: List of employee dictionaries
        all_fields: List of all field names
        
    Returns:
        Dictionary containing summary statistics
    """
    stats = {
        'total_employees': len(employees),
        'total_fields': len(all_fields),
        'total_responses': sum(len(emp) - 2 for emp in employees),  # Exclude 'name' and 'time'
        'field_distribution': {}
    }
    
    # Calculate field distribution
    for field in all_fields:
        non_empty_count = sum(1 for emp in employees if field in emp and emp[field])
        stats['field_distribution'][field] = non_empty_count
    
    return stats


def export_to_csv(employees: List[Dict[str, str]], all_fields: List[str], output_filename: str = 'OUTPUT.csv') -> str:
    """
    Export employee data to CSV format.
    
    Args:
        employees: List of employee dictionaries
        all_fields: List of all field names
        output_filename: Name of the output CSV file
        
    Returns:
        Path to the created CSV file
    """
    # Prepare data for DataFrame
    excel_data = []
    for emp in employees:
        row = {
            'name': emp.get('name', ''),
            'time': emp.get('time', '')
        }
        # Add all fields
        for field in all_fields:
            row[field] = emp.get(field, '')
        excel_data.append(row)
    
    # Create DataFrame and save to CSV (in project root, not lib folder)
    df = pd.DataFrame(excel_data)
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path = os.path.join(project_root, output_filename)
    df.to_csv(output_path, index=False)
    
    print(f"CSV file created: {output_path}")
    return output_path
