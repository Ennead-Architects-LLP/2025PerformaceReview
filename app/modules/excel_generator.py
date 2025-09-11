"""
Excel Report Generator

This module handles the generation of Excel files from employee evaluation data.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict
from .config import Config


def create_excel_output(employees: List[Dict[str, str]], all_fields: List[str], output_filename: str = 'OUTPUT.xlsx') -> str:
    """
    Create Excel file with employee data and professional formatting.
    """
    excel_data = []
    for emp in employees:
        row = {}
        priority_fields = ['employee_name','employee_role','date_of_evaluation','evaluator_name']
        for field in priority_fields:
            if field == 'employee_name':
                row['Employee Name'] = emp.get('employee_name', '')
            else:
                row[field] = emp.get(field, '')
        for field in all_fields:
            if field not in priority_fields:
                row[field] = emp.get(field, '')
        excel_data.append(row)

    df = pandas.DataFrame(excel_data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Evaluations"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    apply_excel_formatting(ws, len(employees), len(df.columns))

    docs_dir = Config.get_docs_dir()
    output_path = Config.get_docs_output_path('xlsx')
    os.makedirs(docs_dir, exist_ok=True)
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
    return output_path


def apply_excel_formatting(ws, num_rows: int, num_cols: int) -> None:
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    data_font = Font(size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = border
            if col == 1:
                cell.alignment = left_alignment
            elif col == 2:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    for col in range(1, num_cols + 1):
        column_letter = ws.cell(row=1, column=col).column_letter
        max_length = 0
        header_length = len(str(ws.cell(row=1, column=col).value))
        max_length = max(max_length, header_length)
        for row in range(2, num_rows + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_length = len(str(cell_value))
                if cell_length > 50:
                    cell_length = 50
                max_length = max(max_length, cell_length)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    for row in range(1, num_rows + 2):
        ws.row_dimensions[row].height = 25

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=num_cols).coordinate}"
